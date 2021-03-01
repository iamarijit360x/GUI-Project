#Arijit Banerjee

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles.borders import Side
from openpyxl.workbook import Workbook
import pathlib
import os
from openpyxl.styles import Color,Fill,PatternFill,Font,Border
from openpyxl.cell import Cell

global loc
loc=os.path.dirname(__file__) +'\studentinfo.xlsx'


def format_cell_width(ws):
    for letter in range(1,ws.max_column):
        max_size = 0
        for cell in ws[get_column_letter(letter)]:
            val_to_check = len(str(cell.value))
            if val_to_check > max_size:
               max_size = val_to_check
        ws.column_dimensions[get_column_letter(letter)].width = max_size + 1



def check_file(loc):
    file = pathlib.Path(loc)
    if not(file.exists()):
        data=["Registration Number","Name",'DOB',"Father's Name","Mother's Name","Gender",'Dept','Domicile','Mother\'s Phone Number','Father\'s Phone Number','Email Id','Present Address','Permanent Address'] #first row
        wb = Workbook()
        ws0 = wb.worksheets[0]
        ws0.title = 'CSE'
        ws1 = wb.create_sheet()
        ws1.title = 'IT'
        ws2 = wb.create_sheet()
        
        ws2.title = 'ECE'
        ws3 = wb.create_sheet()
        
        ws3.title = 'EE'
        ws0.append(data)
        ws1.append(data)
        ws2.append(data)
        ws3.append(data)
        ws = wb.active
        for sheet in wb.sheetnames:
               
                page = wb.get_sheet_by_name(sheet)
                
                for i in range(1,len(data)+1):
                    page.cell(row=1, column=i).font=Font(bold=True)


       
        wb.save(filename =loc)
       

       
def save_xls_file(arr,loc):
    workbook_name =loc
    wb =load_workbook(workbook_name)
    reg='ST169/'
    
    if(arr[5]=="CSE"):
        ch=0
        cl='75C825'
    elif(arr[5]=="IT"):
        ch=1
        cl='41D1FB'
        
    elif(arr[5]=="ECE"):
         ch=2
         cl='FFFC00'
    elif(arr[5]=="EE"):
         ch=3
         cl='F03C60'
    page = wb.get_sheet_by_name(wb.sheetnames[ch])
    reg+=wb.sheetnames[ch]+str(page.max_row)
    arr.insert(0,reg)
    page.append(arr)
    format_cell_width(page)
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for i in range(1,page.max_column+1):
        page.cell(row=page.max_row, column=i).fill = PatternFill(fgColor=cl, fill_type = 'solid')
        page.cell(row=page.max_row, column=i).border =thin_border
       
    wb.save(filename =workbook_name)

    
if __name__ == '__main__':
    
    loc=os.path.dirname(__file__) +'\studentinfo.xlsx'
    check_file(loc)
