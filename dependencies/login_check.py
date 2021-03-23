import os
from tkinter.constants import TRUE
import openpyxl
import pathlib
global loc
loc=os.path.dirname(__file__) +'\studentlogininfo.xlsx'
def check_file(loc):
    file = pathlib.Path(loc)
    if not(file.exists()):
        data=["Email","Password"]
        wb =openpyxl.workbook.Workbook()
        ws0 = wb.worksheets[0]
        ws0.append(data)
        wb.save(filename =loc)

def check_email_present(email):
    check_file(loc)
    book =openpyxl.load_workbook(loc)
    sheet = book.active
    for i in range(1,sheet.max_row+1):
        if(((sheet.cell(row=i,column=1)).value)==email):
            return True
    return False
def load(email):
    loc2=os.path.dirname(__file__) +'\studentinfo.xlsx'
    book =openpyxl.load_workbook(loc2)
    sheet = book.active
    info_dict={}
    for i in range(1,sheet.max_row+1):
        if(((sheet.cell(row=i,column=10)).value)==email):
            for j in range(1,13):
                info_dict[(sheet.cell(row=1,column=j)).value]=(sheet.cell(row=i,column=j)).value
            return info_dict
def creditional_check(email,password):
    check_file(loc)
    book =openpyxl.load_workbook(loc)
    sheet = book.active
    for i in range(1,sheet.max_row+1):
        if(((sheet.cell(row=i,column=1)).value)==email):
            if(sheet.cell(row=i,column=2)==password):
                return load(email)

def save_xls_file(arr):
    workbook_name =loc
    wb =openpyxl.load_workbook(workbook_name)
    page = wb.active
    page.append(arr)
    wb.save(filename =workbook_name)

#arr=['arijit.bandyos7@gmail.com','1234']

